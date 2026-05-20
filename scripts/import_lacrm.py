#!/usr/bin/env python3
"""
Konomocha CRM — LACRM Import Script
Usage: python3 import_lacrm.py /path/to/LessAnnoyingCRM_export.xlsx
Run from: /home/thierry/konomocha-crm (with venv active)
"""

import sys
import os
sys.path.insert(0, '/home/thierry/konomocha-crm')

from dotenv import load_dotenv
load_dotenv()

import openpyxl
from app.database import SessionLocal
from app.models.models import Contact

def clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def import_lacrm(filepath):
    print(f"Opening: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    print(f"Columns: {len(headers)}")

    def col(row, name):
        try:
            idx = headers.index(name)
            return clean(row[idx])
        except (ValueError, IndexError):
            return None

    db = SessionLocal()
    added = 0
    skipped = 0
    errors = 0

    try:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            try:
                # Build name
                first = col(row, 'First Name') or ''
                last  = col(row, 'Last Name')  or ''
                name  = (str(first) + ' ' + str(last)).strip()
                if not name:
                    name = col(row, 'Company Name') or 'Unknown'

                # Build address from primary fields
                parts = [
                    col(row, 'Primary Street 1'),
                    col(row, 'Primary Street 2'),
                    col(row, 'Primary City'),
                    col(row, 'Primary State'),
                    col(row, 'Primary Zip'),
                ]
                address = ', '.join(p for p in parts if p) or None

                # Tags from Groups + type
                groups = col(row, 'Groups') or ''
                rec_type = col(row, 'Type') or ''
                tags_parts = [t for t in [groups, rec_type] if t]
                tags = ', '.join(tags_parts) if tags_parts else None

                # Phone — strip non-digits for storage
                phone_raw = col(row, 'Primary Phone')
                phone = clean(phone_raw)

                # Country — prefer the custom 'Country' field (col 26), fall back to Primary Country
                country = col(row, 'Country') or col(row, 'Primary Country')

                # Notes
                notes = col(row, 'Background Info') or col(row, 'Notes')

                c = Contact(
                    name=name,
                    company=col(row, 'Company Name'),
                    email=col(row, 'Primary Email'),
                    phone=phone,
                    country=country,
                    address=address,
                    tags=tags,
                    notes=notes,
                    source='lacrm_import',
                )
                db.add(c)
                added += 1

                # Commit in batches
                if added % 500 == 0:
                    db.commit()
                    print(f"  {added} contacts imported…")

            except Exception as e:
                errors += 1
                if errors <= 5:
                    print(f"  Row {i} error: {e}")

        db.commit()
        print(f"\nDone. Imported: {added}, Errors: {errors}")

    finally:
        db.close()

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 import_lacrm.py /path/to/export.xlsx")
        sys.exit(1)
    import_lacrm(sys.argv[1])
