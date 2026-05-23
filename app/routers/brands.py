from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from typing import Optional
from app.database import get_db
from app.models.models import Brand, User
from app.auth import get_current_user, require_admin
from app.schemas.brands import BrandCreate

router = APIRouter(prefix="/api/brands", tags=["brands"])

def brand_to_dict(b: Brand):
    return {"id": b.id, "name": b.name, "notes": b.notes, "is_active": b.is_active}

@router.get("")
def list_brands(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    brands = db.query(Brand).order_by(Brand.name).all()
    return [brand_to_dict(b) for b in brands]

@router.post("")
def create_brand(data: BrandCreate, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    b = Brand(**data.dict())
    db.add(b)
    db.commit()
    db.refresh(b)
    return brand_to_dict(b)

@router.put("/{brand_id}")
def update_brand(brand_id: int, data: BrandCreate, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    b = db.query(Brand).filter(Brand.id == brand_id).first()
    if not b:
        raise HTTPException(status_code=404, detail="Brand not found")
    for k, v in data.dict(exclude_unset=True).items():
        setattr(b, k, v)
    db.commit()
    db.refresh(b)
    return brand_to_dict(b)


@router.get("/{brand_id}/report.xlsx")
def brand_report(brand_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    from app.models.models import PipelineEntry, Order, Contact
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    brand = db.query(Brand).filter(Brand.id == brand_id).first()
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")

    pipeline = (db.query(PipelineEntry)
                  .filter(PipelineEntry.brand_id == brand_id)
                  .order_by(PipelineEntry.updated_at.desc())
                  .all())
    orders = (db.query(Order)
                .filter(Order.brand_id == brand_id)
                .order_by(Order.order_date.desc())
                .all())

    wb = openpyxl.Workbook()

    # ── Style helpers ──────────────────────────────────────────────────────
    NAVY   = "002147"
    BLUE   = "76BCE0"
    WHITE  = "FFFFFF"
    LGREY  = "F5F7FA"

    hdr_font  = Font(name="Calibri", bold=True, color=WHITE, size=11)
    hdr_fill  = PatternFill("solid", fgColor=NAVY)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sub_font  = Font(name="Calibri", bold=True, color=NAVY, size=10)
    sub_fill  = PatternFill("solid", fgColor=BLUE)
    sub_align = Alignment(horizontal="center", vertical="center")

    body_font  = Font(name="Calibri", size=10)
    body_align = Alignment(vertical="center")
    alt_fill   = PatternFill("solid", fgColor=LGREY)
    thin_side  = Side(style="thin", color="DDDDDD")
    thin_border= Border(bottom=thin_side)

    def style_header_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font   = hdr_font
            cell.fill   = hdr_fill
            cell.alignment = hdr_align
            cell.border = Border(bottom=Side(style="medium", color=BLUE))

    def style_subheader_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font   = sub_font
            cell.fill   = sub_fill
            cell.alignment = sub_align

    def style_body_row(ws, row, cols, alt=False):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font      = body_font
            cell.alignment = body_align
            cell.border    = thin_border
            if alt:
                cell.fill = alt_fill

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 1: Pipeline ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Pipeline"
    ws1.freeze_panes = "A3"

    # Title row
    ws1.merge_cells("A1:G1")
    t = ws1["A1"]
    t.value     = f"{brand.name} — Pipeline Report"
    t.font      = Font(name="Calibri", bold=True, size=14, color=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 28

    # Column headers
    p_cols = ["Contact", "Company", "Status", "Value (USD)", "Next Action", "Due Date", "Owner"]
    for i, h in enumerate(p_cols, 1):
        ws1.cell(row=2, column=i, value=h)
    style_subheader_row(ws1, 2, len(p_cols))
    ws1.row_dimensions[2].height = 20

    for ri, e in enumerate(pipeline, 3):
        contact_name    = e.contact.name    if e.contact else ""
        contact_company = e.contact.company if e.contact else ""
        row_data = [
            contact_name,
            contact_company,
            e.status,
            float(e.potential_value) if e.potential_value else 0,
            e.next_action or "",
            e.fob_date.strftime("%d %b %Y") if e.fob_date else "",
            e.owner.name if e.owner else "",
        ]
        for ci, val in enumerate(row_data, 1):
            ws1.cell(row=ri, column=ci, value=val)
        style_body_row(ws1, ri, len(p_cols), alt=(ri % 2 == 0))
        # Right-align the value column
        ws1.cell(row=ri, column=4).alignment = Alignment(horizontal="right", vertical="center")
        ws1.cell(row=ri, column=4).number_format = '#,##0.00'

    set_col_widths(ws1, [22, 22, 20, 14, 30, 12, 16])
    style_header_row(ws1, 1, len(p_cols))  # apply navy to merged title

    # ── Sheet 2: Orders ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Orders")
    ws2.freeze_panes = "A3"

    ws2.merge_cells("A1:H1")
    t2 = ws2["A1"]
    t2.value     = f"{brand.name} — Orders Report"
    t2.font      = Font(name="Calibri", bold=True, size=14, color=NAVY)
    t2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 28

    o_cols = ["Order Date", "Contact", "Company", "Value", "Currency", "Status", "Commission %", "Net Commission"]
    for i, h in enumerate(o_cols, 1):
        ws2.cell(row=2, column=i, value=h)
    style_subheader_row(ws2, 2, len(o_cols))
    ws2.row_dimensions[2].height = 20

    for ri, o in enumerate(orders, 3):
        contact_name    = o.contact.name    if o.contact else ""
        contact_company = o.contact.company if o.contact else ""
        row_data = [
            o.order_date.strftime("%d %b %Y") if o.order_date else "",
            contact_name,
            contact_company,
            float(o.order_value) if o.order_value else 0,
            o.currency or "USD",
            o.status or "",
            float(o.gross_commission_rate) if o.gross_commission_rate else 0,
            float(o.net_commission) if o.net_commission else 0,
        ]
        for ci, val in enumerate(row_data, 1):
            ws2.cell(row=ri, column=ci, value=val)
        style_body_row(ws2, ri, len(o_cols), alt=(ri % 2 == 0))
        for col in [4, 8]:
            ws2.cell(row=ri, column=col).alignment = Alignment(horizontal="right", vertical="center")
            ws2.cell(row=ri, column=col).number_format = '#,##0.00'
        ws2.cell(row=ri, column=7).number_format = '0.00%'

    set_col_widths(ws2, [13, 22, 22, 14, 10, 18, 14, 16])
    style_header_row(ws2, 1, len(o_cols))

    # ── Stream the file ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = brand.name.replace(" ", "_") + "_report.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

