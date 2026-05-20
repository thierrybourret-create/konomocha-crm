#!/usr/bin/env python3
"""
Fix Sales Region and Country on existing contacts using the LACRM contacts export.
Run from: /home/thierry/konomocha-crm (with venv active)
Usage: python3 scripts/fix_contact_regions.py scripts/LessAnnoyingCRM_contacts.xlsx
"""
import sys
sys.path.insert(0, '/home/thierry/konomocha-crm')
from dotenv import load_dotenv
load_dotenv()

import openpyxl
from app.database import SessionLocal
from app.models.models import Contact

def clean(val):
    if val is None: return None
    s = str(val).strip()
    return s if s else None

def fix_regions(filepath):
    print(f"Opening: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    def last_col(name):
        indices = [i for i, h in enumerate(headers) if h == name]
        return indices[-1] if indices else None

    sr_idx  = last_col('Sales Region')
    co_idx  = last_col('Country')
    grp_idx = last_col('Groups')
    email_idx = headers.index('Primary Email') if 'Primary Email' in headers else None
    print(f"Sales Region col: {sr_idx}, Country col: {co_idx}")

    db = SessionLocal()
    updated = skipped = 0

    try:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            email = clean(row[email_idx]) if email_idx else None
            if not email:
                skipped += 1
                continue

            contact = db.query(Contact).filter(Contact.email == email).first()
            if not contact:
                skipped += 1
                continue

            country      = clean(row[co_idx])  if co_idx  is not None else None
            sales_region = clean(row[sr_idx])  if sr_idx  is not None else None
            groups       = clean(row[grp_idx]) if grp_idx is not None else None

            if country: contact.country = country

            # Rebuild tags preserving existing, adding region
            existing_tags = [t.strip() for t in (contact.tags or '').split(',') if t.strip()]
            new_tags = list(dict.fromkeys(existing_tags))  # dedupe
            if groups and groups not in new_tags: new_tags.append(groups)
            if sales_region and sales_region not in new_tags: new_tags.append(sales_region)
            # Remove generic 'Contact' tag
            new_tags = [t for t in new_tags if t.lower() not in ('contact',)]
            contact.tags = ', '.join(new_tags)
            updated += 1

            if updated % 500 == 0:
                db.commit()
                print(f"  {updated} updated…")

        db.commit()
        print(f"\nDone. Updated: {updated}, Skipped: {skipped}")

    finally:
        db.close()

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/fix_contact_regions.py /path/to/contacts.xlsx")
        sys.exit(1)
    fix_regions(sys.argv[1])
