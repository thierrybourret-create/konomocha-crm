from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, func
from datetime import datetime, timedelta
from typing import Optional
import io, csv
from app.database import get_db
from app.models.models import Contact, PipelineEntry, Order, EmailLog, User, ContactNote, ContactTask
from app.auth import get_current_user, require_admin

router = APIRouter(prefix="/api/reports", tags=["reports"])

DEF_REGIONS = ['Africa','Asia','Balkans','Eastern Europe','Europe','India','Middle East',
                'North America','Oceania','South America']

DEF_COUNTRIES = [
    'Afghanistan','Albania','Algeria','Argentina','Australia','Austria',
    'Bahrain','Bangladesh','Belgium','Brazil','Bulgaria','Cambodia','Canada',
    'Chile','China','Colombia','Croatia','Cyprus','Czech Republic','Denmark',
    'Ecuador','Egypt','Estonia','Fiji','Finland','France','Georgia','Germany','Ghana',
    'Gibraltar','Greece','Hong Kong','Hungary','India','Indonesia','Iran','Iraq','Ireland',
    'Israel','Italy','Japan','Jordan','Kazakhstan','Kenya','Kuwait','Latvia',
    'Lebanon','Lithuania','Luxembourg','Malaysia','Malta','Mexico','Morocco',
    'Netherlands','New Zealand','Nigeria','Norway','Oman','Pakistan','Philippines',
    'Poland','Portugal','Qatar','Romania','Russia','Saudi Arabia','Singapore',
    'Slovakia','Slovenia','South Africa','South Korea','Spain','Sri Lanka',
    'Angola','Armenia','Azerbaijan','Belarus','Bosnia','Botswana',
    'Costa Rica','Dominican Republic','El Salvador',
    'Iceland','Jamaica','Korea','Liechtenstein',
    'Macedonia','Madagascar','Mauritius','Moldova','Mongolia','Myanmar',
    'Namibia','Panama','Palestine','Paraguay','Peru','Puerto Rico',
    'Senegal','Serbia','Sweden','Switzerland',
    'Taiwan','Thailand','Trinidad and Tobago','Tunisia','Turkey',
    'Ukraine','United Arab Emirates','United Kingdom',
    'United States of America','Uruguay','Venezuela','Vietnam','Zimbabwe',
]

DEF_TAGS = ['Agent','CloseOut','DND','Distributor','Manufacturer','Network','Principal','Retailer','Stationery']

def _has_region(tags):
    if not tags: return False
    return any(r in tags for r in DEF_REGIONS)

def _has_group(tags):
    if not tags: return False
    parts = [t.strip() for t in tags.split(',') if t.strip()]
    return any(t for t in parts if t not in DEF_REGIONS and t not in ('Contact', 'Company'))

def _get_region(tags):
    if not tags: return None
    for r in DEF_REGIONS:
        if r in tags: return r
    return None

def _score(c):
    checks = [
        bool(c.email     and c.email.strip()),
        bool(c.phone     and c.phone.strip()),
        bool(c.job_title and c.job_title.strip()),
        bool(c.country   and c.country.strip()),
        _has_region(c.tags),
        _has_group(c.tags),
    ]
    score = int(sum(checks) / len(checks) * 100)
    missing = []
    labels  = ['Email','Phone','Job Title','Country','Region','Tags']
    for ok, lbl in zip(checks, labels):
        if not ok: missing.append(lbl)
    return score, missing

def _row(c):
    score, missing = _score(c)
    has_region      = _has_region(c.tags)
    has_country     = bool(c.country and c.country.strip())
    wrong_region    = has_country and not has_region
    invalid_country = has_country and c.country.strip() not in DEF_COUNTRIES
    return {
        "id":              c.id,
        "name":            c.name or "",
        "company":         c.company or "",
        "email":           c.email or "",
        "phone":           c.phone or "",
        "job_title":       c.job_title or "",
        "country":         c.country or "",
        "region":          _get_region(c.tags) or "",
        "tags":            c.tags or "",
        "score":           score,
        "missing":         missing,
        "wrong_region":    wrong_region,
        "invalid_country": invalid_country,
        "complete":        score == 100,
        "owner":           c.owner.name if c.owner else "",
    }



def get_db_probabilities(db):
    """Load pipeline probabilities from DB, fall back to constants."""
    from app.models.models import AppStage
    stages = db.query(AppStage).filter(AppStage.stage_type == 'pipeline').all()
    if stages:
        return {s.name: (s.probability or 0) for s in stages}
    from app.constants import PIPELINE_PROBABILITIES
    return PIPELINE_PROBABILITIES

@router.get("/contact-quality")
def contact_quality(
    filter:       str = Query("all"),
    search:       str = Query(""),
    owner_id:     Optional[int] = Query(None),
    contact_type: str = Query("contacts"),   # contacts | companies | all
    page:         int = Query(1, ge=1),
    per_page:     int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
):
    q = db.query(Contact)
    if contact_type == "contacts":
        q = q.filter(Contact.source.notin_(("lacrm_company_import", "pipeline_import")))
    elif contact_type == "companies":
        q = q.filter(Contact.source.in_(("lacrm_company_import", "pipeline_import")))

    if search:
        s = f"%{search}%"
        q = q.filter(or_(Contact.name.ilike(s), Contact.company.ilike(s), Contact.email.ilike(s)))

    if owner_id:
        q = q.filter(Contact.owner_id == owner_id)

    contacts = q.all()
    all_rows = [_row(c) for c in contacts]

    total            = len(all_rows)
    n_missing        = sum(1 for r in all_rows if not r["complete"])
    n_complete       = sum(1 for r in all_rows if r["complete"])
    n_wrong_rgn      = sum(1 for r in all_rows if r["wrong_region"])
    n_invalid_ctry   = sum(1 for r in all_rows if r["invalid_country"])

    if filter == "missing":
        filtered = [r for r in all_rows if not r["complete"]]
    elif filter == "complete":
        filtered = [r for r in all_rows if r["complete"]]
    elif filter == "wrong_region":
        filtered = [r for r in all_rows if r["wrong_region"]]
    elif filter == "invalid_country":
        filtered = [r for r in all_rows if r["invalid_country"]]
    else:
        filtered = all_rows

    filtered.sort(key=lambda r: r["score"])
    start     = (page - 1) * per_page
    page_rows = filtered[start:start + per_page]

    return {
        "total":             total,
        "n_missing":         n_missing,
        "n_complete":        n_complete,
        "n_wrong_region":    n_wrong_rgn,
        "n_invalid_country": n_invalid_ctry,
        "filtered_total":    len(filtered),
        "page":              page,
        "per_page":          per_page,
        "results":           page_rows,
    }

@router.get("/contact-quality-csv")
def contact_quality_csv(
    filter:       str = Query("all"),
    search:       str = Query(""),
    owner_id:     Optional[int] = Query(None),
    contact_type: str = Query("contacts"),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
):
    q = db.query(Contact)
    if contact_type == "contacts":
        q = q.filter(Contact.source.notin_(("lacrm_company_import", "pipeline_import")))
    elif contact_type == "companies":
        q = q.filter(Contact.source.in_(("lacrm_company_import", "pipeline_import")))

    if search:
        s = f"%{search}%"
        q = q.filter(or_(Contact.name.ilike(s), Contact.company.ilike(s), Contact.email.ilike(s)))

    if owner_id:
        q = q.filter(Contact.owner_id == owner_id)

    contacts = q.all()
    all_rows = [_row(c) for c in contacts]

    if filter == "missing":
        rows = [r for r in all_rows if not r["complete"]]
    elif filter == "complete":
        rows = [r for r in all_rows if r["complete"]]
    elif filter == "wrong_region":
        rows = [r for r in all_rows if r["wrong_region"]]
    elif filter == "invalid_country":
        rows = [r for r in all_rows if r["invalid_country"]]
    else:
        rows = all_rows

    rows.sort(key=lambda r: r["score"])

    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["Name","Company","Job Title","Email","Phone","Country","Region","Owner","Score","Missing Fields","Status"])
    for r in rows:
        w.writerow([
            r["name"], r["company"], r["job_title"], r["email"], r["phone"],
            r["country"], r["region"], r["owner"], f"{r['score']}%",
            ", ".join(r["missing"]),
            "Complete" if r["complete"] else "Missing fields",
        ])
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=contact_quality.csv"},
    )

@router.get("/activity")
def activity_report(
    period: str = Query("30d"),
    owner_id: int = Query(None),
    db: Session = Depends(get_db),
    current_user = Depends(require_admin),
):
    from datetime import timezone as _tz
    HK = _tz(timedelta(hours=8))
    now_hk = datetime.now(HK)
    deltas = {"1d": 0, "7d": 6, "30d": 29, "90d": 89, "365d": 364}
    days_back = deltas.get(period, 29)
    # Start of day N days ago in HK time, converted to UTC naive for DB
    start_hk = (now_hk - timedelta(days=days_back)).replace(hour=0, minute=0, second=0, microsecond=0)
    start = start_hk.astimezone(_tz.utc).replace(tzinfo=None)

    def _c(q): return q.count()

    cq  = db.query(Contact)
    pq  = db.query(PipelineEntry)
    oq  = db.query(Order)
    eiq = db.query(EmailLog).filter(EmailLog.direction == "inbound")
    eoq = db.query(EmailLog).filter(EmailLog.direction == "outbound")
    tq  = db.query(ContactTask)
    nq  = db.query(ContactNote)

    if owner_id:
        cq  = cq.filter(Contact.owner_id == owner_id)
        pq  = pq.filter(PipelineEntry.owner_id == owner_id)
        eoq = eoq.filter(EmailLog.logged_by_id == owner_id)
        tq  = tq.filter(ContactTask.assigned_to_id == owner_id)
        nq  = nq.filter(ContactNote.author_id == owner_id)
        # orders and inbound emails have no direct owner

    return {
        "period":               period,
        "since":                start_hk.date().isoformat(),
        "new_contacts":         _c(cq.filter(Contact.source.notin_(("lacrm_company_import", "pipeline_import")), Contact.created_at >= start)),
        "new_companies":        _c(cq.filter(Contact.source.in_(("lacrm_company_import", "pipeline_import")), Contact.created_at >= start)),
        "new_pipeline_entries": _c(pq.filter(PipelineEntry.created_at >= start)),
        "new_orders":           _c(oq.filter(Order.created_at >= start)),
        "emails_inbound":       _c(eiq.filter(EmailLog.sent_at >= start)),
        "emails_outbound":      _c(eoq.filter(EmailLog.sent_at >= start)),
        "new_tasks":            _c(tq.filter(ContactTask.created_at >= start)),
        "completed_tasks":      _c(tq.filter(ContactTask.completed == True, ContactTask.completed_at >= start)),
        "new_notes":            _c(nq.filter(ContactNote.created_at >= start)),
    }


@router.get("/tasks")
def task_report(
    filter: str = Query("open"),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
):
    from app.models.models import ContactTask
    from datetime import date
    q = db.query(ContactTask)
    today = date.today()
    if filter == "open":
        q = q.filter(ContactTask.completed == False)
    elif filter == "overdue":
        q = q.filter(ContactTask.completed == False, ContactTask.due_date < today)
    elif filter == "completed":
        q = q.filter(ContactTask.completed == True)
    tasks = q.order_by(ContactTask.due_date.asc().nullslast(), ContactTask.created_at.desc()).all()
    return [{"id": t.id, "title": t.title,
             "due_date": t.due_date.isoformat() if t.due_date else None,
             "completed": t.completed,
             "contact_id": t.contact_id,
             "contact_name": t.contact.name if t.contact else None,
             "contact_company": t.contact.company if t.contact else None,
             "assigned_to": t.assigned_to.name if t.assigned_to else None,
             "created_by": t.created_by.name if t.created_by else None,
             "created_at": t.created_at.isoformat()} for t in tasks]


@router.get("/principal-report.xlsx")
def principal_report(
    brands: str = Query(..., description="Comma-separated brand IDs"),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """Download a combined Excel report for selected brands."""
    from app.models.models import Brand, PipelineEntry, Order
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    try:
        brand_ids = [int(x.strip()) for x in brands.split(",") if x.strip()]
    except ValueError:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="Invalid brand IDs")

    selected_brands = db.query(Brand).filter(Brand.id.in_(brand_ids)).order_by(Brand.name).all()
    if not selected_brands:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="No brands found")

    pipeline_rows = (db.query(PipelineEntry)
                       .filter(PipelineEntry.brand_id.in_(brand_ids))
                       .order_by(PipelineEntry.brand_id, PipelineEntry.updated_at.desc())
                       .all())
    order_rows = (db.query(Order)
                    .filter(Order.brand_id.in_(brand_ids))
                    .order_by(Order.brand_id, Order.order_date.desc())
                    .all())

    # ── Style helpers ──────────────────────────────────────────────────
    NAVY  = "002147"
    BLUE  = "76BCE0"
    WHITE = "FFFFFF"
    LGREY = "F5F7FA"
    hf = Font(name="Calibri", bold=True, color=WHITE, size=11)
    hfill = PatternFill("solid", fgColor=NAVY)
    ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sf = Font(name="Calibri", bold=True, color=NAVY, size=10)
    sfill = PatternFill("solid", fgColor=BLUE)
    sa = Alignment(horizontal="center", vertical="center")
    bf = Font(name="Calibri", size=10)
    ba = Alignment(vertical="center")
    altfill = PatternFill("solid", fgColor=LGREY)
    thin = Side(style="thin", color="DDDDDD")
    tborder = Border(bottom=thin)
    num_fmt = '#,##0.00'

    def hdr(ws, row, ncols):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.font=hf; cell.fill=hfill; cell.alignment=ha
            cell.border=Border(bottom=Side(style="medium",color=BLUE))

    def subhdr(ws, row, ncols):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.font=sf; cell.fill=sfill; cell.alignment=sa

    def bodyrow(ws, row, ncols, alt=False):
        for c in range(1, ncols+1):
            cell = ws.cell(row=row, column=c)
            cell.font=bf; cell.alignment=ba; cell.border=tborder
            if alt: cell.fill=altfill

    def widths(ws, ws_list):
        for i, w in enumerate(ws_list, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = openpyxl.Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────
    ws0 = wb.active; ws0.title = "Summary"
    ws0["A1"] = "Principal Report — Konomocha (HK)"
    ws0["A1"].font = Font(name="Calibri", bold=True, size=16, color=NAVY)
    ws0.merge_cells("A1:D1"); ws0.row_dimensions[1].height = 32

    from datetime import date as _date
    ws0["A2"] = f"Generated: {_date.today().strftime('%d %b %Y')}"
    ws0["A2"].font = Font(name="Calibri", size=10, color="888888")
    ws0.row_dimensions[2].height = 16
    ws0.row_dimensions[3].height = 8

    summary_headers = ["Brand", "Active Pipeline Deals", "Total Orders", "Total Order Value (USD)"]
    for i, h in enumerate(summary_headers, 1):
        ws0.cell(row=4, column=i, value=h)
    subhdr(ws0, 4, len(summary_headers)); ws0.row_dimensions[4].height = 20

    for ri, brand in enumerate(selected_brands, 5):
        bp = [e for e in pipeline_rows if e.brand_id == brand.id and e.status not in ("Closed / No Action","Cancelled")]
        bo = [o for o in order_rows  if o.brand_id == brand.id]
        row_data = [
            brand.name,
            len(bp),
            len(bo),
            sum(float(o.order_value)     for o in bo if o.order_value),
        ]
        for ci, val in enumerate(row_data, 1):
            ws0.cell(row=ri, column=ci, value=val)
        bodyrow(ws0, ri, len(summary_headers), alt=(ri%2==0))
        ws0.cell(row=ri, column=4).number_format = num_fmt
        ws0.cell(row=ri, column=4).alignment = Alignment(horizontal="right", vertical="center")
        ws0.cell(row=ri, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws0.cell(row=ri, column=3).alignment = Alignment(horizontal="center", vertical="center")
    widths(ws0, [24, 22, 14, 24])
    hdr(ws0, 1, len(summary_headers))

    # ── Pipeline sheet ─────────────────────────────────────────────────
    ws1 = wb.create_sheet("Pipeline"); ws1.freeze_panes = "A3"
    ws1.merge_cells("A1:G1")
    t = ws1["A1"]; t.value = "Pipeline — All Selected Brands"
    t.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    t.alignment = Alignment(horizontal="left", vertical="center"); ws1.row_dimensions[1].height = 28
    p_cols = ["Brand", "Contact", "Company", "Status", "Next Action", "Due Date", "Owner"]
    for i, h in enumerate(p_cols, 1): ws1.cell(row=2, column=i, value=h)
    subhdr(ws1, 2, len(p_cols)); ws1.row_dimensions[2].height = 20
    for ri, e in enumerate(pipeline_rows, 3):
        row_data = [
            e.brand.name if e.brand else "",
            e.contact.name    if e.contact else "",
            e.contact.company if e.contact else "",
            e.status,
            e.next_action or "",
            e.fob_date.strftime("%d %b %Y") if e.fob_date else "",
            e.owner.name if e.owner else "",
        ]
        for ci, val in enumerate(row_data, 1): ws1.cell(row=ri, column=ci, value=val)
        bodyrow(ws1, ri, len(p_cols), alt=(ri%2==0))
    widths(ws1, [18, 22, 22, 20, 30, 12, 16])
    hdr(ws1, 1, len(p_cols))

    # ── Orders sheet ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Orders"); ws2.freeze_panes = "A3"
    ws2.merge_cells("A1:I1")
    t2 = ws2["A1"]; t2.value = "Orders — All Selected Brands"
    t2.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    t2.alignment = Alignment(horizontal="left", vertical="center"); ws2.row_dimensions[1].height = 28
    o_cols = ["Brand", "Order Date", "Contact", "Company", "Value", "Currency", "Status", "Commission %", "Net Commission"]
    for i, h in enumerate(o_cols, 1): ws2.cell(row=2, column=i, value=h)
    subhdr(ws2, 2, len(o_cols)); ws2.row_dimensions[2].height = 20
    for ri, o in enumerate(order_rows, 3):
        row_data = [
            o.brand.name if o.brand else "",
            o.order_date.strftime("%d %b %Y") if o.order_date else "",
            o.contact.name    if o.contact else "",
            o.contact.company if o.contact else "",
            float(o.order_value) if o.order_value else 0,
            o.currency or "USD",
            o.status or "",
            float(o.gross_commission_rate) if o.gross_commission_rate else 0,
            float(o.net_commission) if o.net_commission else 0,
        ]
        for ci, val in enumerate(row_data, 1): ws2.cell(row=ri, column=ci, value=val)
        bodyrow(ws2, ri, len(o_cols), alt=(ri%2==0))
        for col in [5, 9]:
            ws2.cell(row=ri, column=col).number_format = num_fmt
            ws2.cell(row=ri, column=col).alignment = Alignment(horizontal="right", vertical="center")
    widths(ws2, [18, 13, 22, 22, 14, 10, 18, 14, 16])
    hdr(ws2, 1, len(o_cols))

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="principal_report.xlsx"'}
    )


@router.get("/commission-forecast")
def commission_forecast(db: Session = Depends(get_db), current_user=Depends(require_admin)):
    from datetime import date as _date
    from app.models.models import Order, PipelineEntry
    from app.constants import PIPELINE_PROBABILITIES, COMMISSION_LAG_DAYS, fy_label, add_months
    today = _date.today()
    # Generate 24 months: Apr 2026 to Mar 2028
    start = _date(2026, 4, 1)
    months_data = []
    for i in range(24):
        m_start = add_months(start, i)
        m_end = add_months(start, i + 1)
        label = m_start.strftime("%b %Y")
        month_key = m_start.strftime("%Y-%m")
        # Actual: commission_paid orders where commission_paid_date in this month
        actual = db.query(func.sum(Order.net_commission)).filter(
            Order.status == 'commission_paid',
            Order.commission_paid_date >= m_start,
            Order.commission_paid_date < m_end,
        ).scalar() or 0
        # Confirmed: shipped (not commission_paid) where ship_date+30 in this month
        conf_orders = db.query(Order).filter(
            Order.status.in_(['shipped', 'fully_paid', 'commission_invoiced']),
            Order.ship_date.isnot(None),
        ).all()
        confirmed = sum(
            float(o.net_commission or 0) for o in conf_orders
            if o.ship_date and m_start <= (o.ship_date + timedelta(days=COMMISSION_LAG_DAYS)) < m_end
        )
        # Weighted pipeline: fob_date+30 in this month, prob > 0
        pipe_entries = db.query(PipelineEntry).filter(PipelineEntry.fob_date.isnot(None)).all()
        weighted = sum(
            float(e.potential_value or 0) * get_db_probabilities(db).get(e.status, 0) / 100
            for e in pipe_entries
            if e.fob_date and m_start <= (e.fob_date + timedelta(days=COMMISSION_LAG_DAYS)) < m_end
              and get_db_probabilities(db).get(e.status, 0) > 0
        )
        months_data.append({
            "month":            month_key,
            "label":            label,
            "actual_received":  round(float(actual), 2),
            "confirmed":        round(confirmed, 2),
            "weighted_pipeline":round(weighted, 2),
            "total_forecast":   round(confirmed + weighted, 2),
            "is_past":          m_end <= today,
            "is_current":       m_start <= today < m_end,
        })

    def fy_sum(months, fy_start_year):
        fy_months = [m for m in months if m["month"] >= f"{fy_start_year}-04" and m["month"] < f"{fy_start_year+1}-04"]
        return {
            "label":             fy_label(fy_start_year),
            "actual_received":   round(sum(m["actual_received"]   for m in fy_months), 2),
            "confirmed":         round(sum(m["confirmed"]         for m in fy_months), 2),
            "weighted_pipeline": round(sum(m["weighted_pipeline"] for m in fy_months), 2),
            "total_forecast":    round(sum(m["total_forecast"]    for m in fy_months), 2),
        }

    return {
        "months":     months_data,
        "fy_current": fy_sum(months_data, 2026),
        "fy_next":    fy_sum(months_data, 2027),
    }


@router.get("/bonus")
def bonus_report(
    fy: int = Query(..., description="FY start year, e.g. 2026 for FY2026-27"),
    quarter: str = Query(..., description="Q1, Q2, Q3, or Q4"),
    db: Session = Depends(get_db),
    current_user=Depends(require_admin),
):
    from datetime import date as _date
    from app.models.models import Order, User
    from app.constants import quarter_date_range, quarter_label
    q_num = int(quarter[1]) if quarter and len(quarter) >= 2 else 1
    q_start, q_end = quarter_date_range(fy, q_num)
    agents = db.query(User).filter(User.role == "agent", User.is_active == True).all()
    result_agents = []
    grand_total = 0.0
    for agent in agents:
        orders = db.query(Order).options(
            joinedload(Order.contact), joinedload(Order.brand)
        ).filter(
            Order.owner_id == agent.id,
            Order.status == "commission_paid",
            Order.commission_paid_date >= q_start,
            Order.commission_paid_date < q_end,
        ).all()
        agent_orders = []
        total_net = 0.0
        total_bonus = 0.0
        for o in orders:
            net = float(o.net_commission or 0)
            bonus = float(o.bonus_amount or 0)
            agent_orders.append({
                "order_id":            o.id,
                "contact_name":        o.contact.name if o.contact else None,
                "brand_name":          o.brand.name   if o.brand   else None,
                "net_commission":      net,
                "bonus_amount":        bonus,
                "commission_paid_date":o.commission_paid_date.isoformat() if o.commission_paid_date else None,
                "bonus_paid":          o.bonus_paid_date is not None,
            })
            total_net += net
            total_bonus += bonus
        grand_total += total_bonus
        result_agents.append({
            "owner_id":             agent.id,
            "owner_name":           agent.name,
            "orders":               agent_orders,
            "total_net_commission": round(total_net, 2),
            "total_bonus":          round(total_bonus, 2),
        })
    return {
        "quarter":           quarter_label(fy, q_num),
        "agents":            result_agents,
        "grand_total_bonus": round(grand_total, 2),
    }


@router.get("/my-bonus")
def my_bonus(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    from datetime import date as _date
    from app.models.models import Order, PipelineEntry
    from app.constants import quarter_of, quarter_label, quarter_date_range, PIPELINE_PROBABILITIES, COMMISSION_LAG_DAYS
    if current_user.role == "admin":
        return None
    today = _date.today()
    fsy, q = quarter_of(today)
    q_start, q_end = quarter_date_range(fsy, q)
    # Next quarter
    nq = q + 1 if q < 4 else 1
    nfy = fsy if q < 4 else fsy + 1
    nq_start, nq_end = quarter_date_range(nfy, nq)
    VISIBLE_STATUSES = ['shipped', 'fully_paid', 'commission_invoiced', 'commission_paid', 'bonus_paid']
    orders = db.query(Order).options(
        joinedload(Order.contact), joinedload(Order.brand)
    ).filter(
        Order.owner_id == current_user.id,
        Order.status.in_(VISIBLE_STATUSES),
    ).all()
    order_list = []
    total_earned = total_paid = total_pending = 0.0
    for o in orders:
        net = float(o.net_commission or 0)
        bonus = float(o.bonus_amount or 0)
        order_list.append({
            "order_id":       o.id,
            "contact_name":   o.contact.name if o.contact else None,
            "brand_name":     o.brand.name   if o.brand   else None,
            "net_commission": net,
            "bonus_amount":   bonus,
            "status":         o.status,
            "bonus_paid":     o.bonus_paid_date is not None,
        })
        if o.status in ('commission_paid', 'bonus_paid'):
            total_earned += bonus
        if o.bonus_paid_date:
            total_paid += bonus
        elif o.status in ('shipped', 'fully_paid', 'commission_invoiced'):
            total_pending += bonus
    # Next quarter pipeline projection
    pipe = db.query(PipelineEntry).filter(
        PipelineEntry.owner_id == current_user.id,
        PipelineEntry.fob_date.isnot(None),
    ).all()
    nq_pipeline = sum(
        float(e.potential_value or 0) * get_db_probabilities(db).get(e.status, 0) / 100 * 0.05
        for e in pipe
        if e.fob_date and nq_start <= (e.fob_date + timedelta(days=COMMISSION_LAG_DAYS)) < nq_end
    )
    return {
        "current_quarter": quarter_label(fsy, q),
        "orders":          order_list,
        "summary": {
            "total_earned":  round(total_earned, 2),
            "total_paid":    round(total_paid, 2),
            "total_pending": round(total_pending, 2),
        },
        "next_quarter_pipeline": round(nq_pipeline, 2),
    }

