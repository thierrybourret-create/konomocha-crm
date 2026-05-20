#!/usr/bin/env python3
"""
Konomocha CRM — Companies Import Script (fixes Sales Region & Country)
Imports company records from LACRM companies export.
Run from: /home/thierry/konomocha-crm (with venv active)
Usage: python3 scripts/import_companies.py scripts/LessAnnoyingCRM_companies.xlsx
"""
import sys
sys.path.insert(0, '/home/thierry/konomocha-crm')
from dotenv import load_dotenv
load_dotenv()

import openpyxl
from app.database import SessionLocal
from app.models.models import Contact

def clean(val):
    if val is None: return None
    s = str(val).strip()
    return s if s else None

def import_companies(filepath):
    print(f"Opening: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

    # Use last occurrence of duplicate columns (company data is in second set)
    def last_col(name):
        indices = [i for i, h in enumerate(headers) if h == name]
        return indices[-1] if indices else None

    sr_idx  = last_col('Sales Region')   # 37
    co_idx  = last_col('Country')        # 36
    grp_idx = last_col('Groups')         # 45
    print(f"Sales Region col: {sr_idx}, Country col: {co_idx}, Groups col: {grp_idx}")

    db = SessionLocal()
    added = updated = skipped = errors = 0

    try:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                company_name = clean(row[7])  # Company Name
                if not company_name:
                    skipped += 1
                    continue

                country      = clean(row[co_idx])  if co_idx  is not None else None
                sales_region = clean(row[sr_idx])  if sr_idx  is not None else None
                groups       = clean(row[grp_idx]) if grp_idx is not None else None
                email        = clean(row[8])
                phone        = clean(row[9])
                notes        = clean(row[47])  # Notes

                # Build tags: groups + type
                tags_parts = [t for t in [groups, 'Company'] if t]
                tags = ', '.join(tags_parts)

                # Store sales region in tags so it's searchable
                if sales_region and sales_region not in tags:
                    tags = tags + ', ' + sales_region if tags else sales_region

                # Check if contact with this company name already exists
                existing = db.query(Contact).filter(
                    Contact.company == company_name,
                    Contact.name == company_name
                ).first()

                if existing:
                    # Update country and sales region
                    if country: existing.country = country
                    if tags: existing.tags = tags
                    if email and not existing.email: existing.email = email
                    if phone and not existing.phone: existing.phone = phone
                    if notes and not existing.notes: existing.notes = notes
                    updated += 1
                else:
                    c = Contact(
                        name=company_name,
                        company=company_name,
                        email=email,
                        phone=phone,
                        country=country,
                        tags=tags,
                        notes=notes,
                        source='lacrm_company_import',
                    )
                    db.add(c)
                    added += 1

                if (added + updated) % 500 == 0:
                    db.commit()
                    print(f"  {added} added, {updated} updated…")

            except Exception as e:
                errors += 1
                if errors <= 5:
                    print(f"  Row {i} error: {e}")

        db.commit()
        print(f"\nDone. Added: {added}, Updated: {updated}, Skipped: {skipped}, Errors: {errors}")

    finally:
        db.close()

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/import_companies.py /path/to/companies.xlsx")
        sys.exit(1)
    import_companies(sys.argv[1])
