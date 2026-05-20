#!/usr/bin/env python3
"""
Konomocha CRM — Pipeline Import Script
Imports from konomocha_pipeline_tracker.xlsx
Usage: python3 import_pipeline.py /path/to/pipeline.xlsx
Run from: /home/thierry/konomocha-crm (with venv active)
"""

import sys
import os
from datetime import datetime, date
sys.path.insert(0, '/home/thierry/konomocha-crm')

from dotenv import load_dotenv
load_dotenv()

import openpyxl
from app.database import SessionLocal
from app.models.models import Contact, Brand, PipelineEntry, User, PipelineStatus

# Excel serial date to Python date
def excel_date(serial):
    if not serial or not str(serial).replace('.','').isdigit():
        return None
    try:
        n = int(float(serial))
        if n < 1:
            return None
        from datetime import timedelta
        base = date(1899, 12, 30)
        return base + timedelta(days=n)
    except:
        return None

# Map Excel status values to our enum
STATUS_MAP = {
    'Awaiting Feedback':   PipelineStatus.awaiting_feedback,
    'Awaiting Info':       PipelineStatus.awaiting_info,
    'Awaiting Samples':    PipelineStatus.awaiting_samples,
    'Cancelled':           PipelineStatus.cancelled,
    'Catalogue Sent':      PipelineStatus.catalogue_sent,
    'Closed / No Action':  PipelineStatus.closed,
    'Deposit Paid':        PipelineStatus.deposit_paid,
    'Form Completed':      PipelineStatus.form_completed,
    'In Progress':         PipelineStatus.in_progress,
    'On Hold':             PipelineStatus.on_hold,
    'Order Placed':        PipelineStatus.order_placed,
    'Price List Sent':     PipelineStatus.price_list_sent,
    'Pricing Sent':        PipelineStatus.pricing_sent,
    'Quotation Sent':      PipelineStatus.quotation_sent,
    'Samples Delivered':   PipelineStatus.samples_delivered,
    'Samples Requested':   PipelineStatus.samples_requested,
    'Samples Sent':        PipelineStatus.samples_sent,
    'Stalled':             PipelineStatus.stalled,
    'Cancelled':           PipelineStatus.cancelled,
}

OWNER_MAP = {
    'Thierry': 'thierry@konomocha.com',
    'Zeal':    'zeal@konomocha.com',
    'Erna':    'erna@konomocha.com',
}

def import_pipeline(filepath):
    print(f"Opening: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb['Pipeline']

    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    print(f"Columns: {headers}")

    def col(row, name):
        try:
            idx = headers.index(name)
            v = row[idx]
            return str(v).strip() if v is not None else None
        except (ValueError, IndexError):
            return None

    db = SessionLocal()

    # Cache users
    users = {u.name: u for u in db.query(User).all()}
    print(f"Users: {list(users.keys())}")

    # Cache/create brands
    brand_cache = {b.name: b for b in db.query(Brand).all()}

    def get_or_create_brand(name):
        if not name or name.strip() == '':
            return None
        name = name.strip()
        if name not in brand_cache:
            b = Brand(name=name)
            db.add(b)
            db.flush()
            brand_cache[name] = b
        return brand_cache[name]

    # Cache/create contacts by company name
    contact_cache = {}
    for c in db.query(Contact).all():
        key = (c.name or '').lower()
        contact_cache[key] = c

    def get_or_create_contact(company_name):
        if not company_name:
            return None
        key = company_name.strip().lower()
        if key not in contact_cache:
            c = Contact(name=company_name.strip(), company=company_name.strip(), source='pipeline_import')
            db.add(c)
            db.flush()
            contact_cache[key] = c
        return contact_cache[key]

    added = 0
    skipped = 0
    errors = 0

    # Default owner = Thierry
    default_owner = users.get('Thierry')

    try:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                company_raw = col(row, 'Contact / Company')
                brand_raw   = col(row, 'Brand')
                status_raw  = col(row, 'Status')
                owner_raw   = col(row, 'Owner')
                next_action = col(row, 'Next Action')
                due_raw     = col(row, 'Due Date')
                notes       = col(row, 'Notes')
                updated_raw = col(row, 'Last Updated')

                # Skip blank rows
                if not company_raw and not brand_raw:
                    skipped += 1
                    continue

                contact = get_or_create_contact(company_raw)
                brand   = get_or_create_brand(brand_raw)

                if not contact or not brand:
                    skipped += 1
                    continue

                # Status
                status = STATUS_MAP.get(status_raw, PipelineStatus.in_progress)

                # Owner
                owner_name = (owner_raw or '').strip()
                owner = users.get(owner_name) or default_owner

                # Dates
                due_date = excel_date(due_raw)

                # Updated at
                updated_at = None
                if updated_raw:
                    try:
                        d = excel_date(updated_raw)
                        if d:
                            updated_at = datetime.combine(d, datetime.min.time())
                    except:
                        pass

                e = PipelineEntry(
                    contact_id=contact.id,
                    brand_id=brand.id,
                    status=status,
                    potential_value=0,  # Was blank in Excel — set to 0, update manually
                    next_action=next_action,
                    due_date=due_date,
                    owner_id=owner.id if owner else default_owner.id,
                    notes=notes,
                )
                if updated_at:
                    e.updated_at = updated_at
                    e.created_at = updated_at

                db.add(e)
                added += 1

                if added % 100 == 0:
                    db.commit()
                    print(f"  {added} pipeline entries imported…")

            except Exception as ex:
                errors += 1
                if errors <= 10:
                    print(f"  Row {i} error: {ex}")

        db.commit()
        print(f"\nDone. Imported: {added}, Skipped: {skipped}, Errors: {errors}")
        print("Note: Potential Value set to 0 for all entries — update in the CRM.")

    finally:
        db.close()

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 import_pipeline.py /path/to/pipeline.xlsx")
        sys.exit(1)
    import_pipeline(sys.argv[1])
